# survey/admin.py
from django.contrib import admin
from django.utils.html import format_html
from django.urls import reverse
from django.template.response import TemplateResponse
from django.http import HttpResponseRedirect, HttpResponse
from django.utils.translation import gettext_lazy as _
from django.db.models import Q, Count
from django import forms
from django.shortcuts import render, redirect
import csv
import os
import re
import random
import string
import uuid

from .models import Survey, Question, Response, Answer, QRCode, Option, SurveyQuestion, Category


# ==================== 自定义AdminSite类 ====================

class CustomAdminSite(admin.AdminSite):
    """自定义Admin站点"""
    site_header = _('问卷调查管理系统')
    site_title = _('问卷调查后台')
    index_title = _('系统管理')
    
    def get_app_list(self, request, app_label=None):
        """重写应用列表排序，将问卷应用放在最前面"""
        # 获取默认的应用列表
        app_list = super().get_app_list(request, app_label)
        
        # 创建一个新的应用列表，将'问卷'应用移到最前面
        new_app_list = []
        other_apps = []
        
        for app in app_list:
            if app['name'] == _('问卷'):
                new_app_list.append(app)
            else:
                other_apps.append(app)
        
        # 将其他应用添加到后面
        new_app_list.extend(other_apps)
        
        return new_app_list


# 替换默认的admin.site
admin.site.__class__ = CustomAdminSite


# ==================== 内联表单类 ====================

class OptionInline(admin.TabularInline):
    """问题选项内联表单"""
    model = Option
    extra = 1
    ordering = ['order']
    fields = ['value', 'label', 'order']
    verbose_name = '选项'
    verbose_name_plural = '选项'
    
    def formfield_for_dbfield(self, db_field, **kwargs):
        """优化表单字段显示"""
        field = super().formfield_for_dbfield(db_field, **kwargs)
        if db_field.name == 'label':
            field.widget.attrs['placeholder'] = '请输入选项显示文本'
        elif db_field.name == 'value':
            field.widget.attrs['placeholder'] = '自动生成，可自定义'
        return field


class SurveyQuestionInline(admin.TabularInline):
    """问卷问题关联内联表单"""
    model = SurveyQuestion
    extra = 1
    ordering = ['order']
    fields = ['question', 'order', 'is_required']
    verbose_name = '问卷问题'
    verbose_name_plural = '问卷问题'
    
    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        """优化问题选择框"""
        if db_field.name == 'question':
            # 预加载相关数据以提高性能，只显示激活分类的问题
            kwargs['queryset'] = Question.objects.filter(
                (Q(is_public=True) | Q(created_by=request.user)),
                (Q(category__is_active=True) | Q(category__isnull=True))
            ).select_related('category').order_by('category__name', 'text')
        
        # 先调用父类方法获取表单字段
        form_field = super().formfield_for_foreignkey(db_field, request, **kwargs)
        
        # 然后设置自定义的label_from_instance
        if db_field.name == 'question':
            def label_from_instance(obj):
                category_name = obj.category.name if obj.category else '未分类'
                type_display = obj.get_question_type_display()
                text_preview = obj.text[:50] + '...' if len(obj.text) > 50 else obj.text
                return f"[{category_name}] {text_preview} ({type_display})"
            
            form_field.label_from_instance = label_from_instance
        
        return form_field


# ==================== 模型管理类 ====================

@admin.register(Category)
class CategoryAdmin(admin.ModelAdmin):
    """问题分类管理"""
    list_display = ['name', 'slug', 'is_active', 'created_at', 'question_count']
    list_filter = ['is_active', 'created_at']
    search_fields = ['name', 'description']
    prepopulated_fields = {'slug': ('name',)}
    fields = ['name', 'slug', 'description', 'is_active']
    ordering = ['-created_at']
    actions = ['make_active', 'make_inactive']
    
    def get_queryset(self, request):
        """优化查询集，预加载相关问题计数"""
        return super().get_queryset(request).annotate(
            _question_count=Count('questions')
        )
    
    def question_count(self, obj):
        """统计该分类下的问题数量"""
        return obj._question_count if hasattr(obj, '_question_count') else obj.questions.count()
    question_count.short_description = '问题数量'
    question_count.admin_order_field = '_question_count'
    
    def make_active(self, request, queryset):
        """批量激活选中的分类"""
        queryset.update(is_active=True)
    make_active.short_description = '批量激活分类'
    
    def make_inactive(self, request, queryset):
        """批量停用选中的分类"""
        queryset.update(is_active=False)
    make_inactive.short_description = '批量停用分类'


@admin.register(Question)
class QuestionAdmin(admin.ModelAdmin):
    """问题库管理"""
    list_display = ['text_preview', 'question_type_display', 'category', 'created_by', 
                    'is_public', 'created_at', 'option_count', 'survey_usage_count']
    list_filter = ['question_type', 'category', 'is_public', 'created_at']
    search_fields = ['text', 'category__name']
    # 优化查询当列出 Question 对象时，会同时使用 JOIN 查询预先加载关联的 category 和 created_by 对象
    list_select_related = ['category', 'created_by']  
    ordering = ['category__name', '-created_at']
    inlines = [OptionInline]
    fields = ['text', 'question_type', 'category', 'created_by', 'is_public']
    # 指定模板
    change_list_template = 'admin/survey/question/change_list.html'
    # actions: 定义可用的批量操作。
    actions = ['make_public', 'make_private', 'change_category', 'export_questions', 'export_questions_excel']
    list_editable = ['is_public', 'category']
    list_per_page = 20
    
    # 自定义表单字段
    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == 'created_by' and not request.user.is_superuser:
            kwargs['initial'] = request.user.id
            kwargs['disabled'] = True
        return super().formfield_for_foreignkey(db_field, request, **kwargs)
    
    # 自定义查询集
    # 实现了基于权限的数据隔离，普通用户只能看到公开的或自己创建的问题。
    def get_queryset(self, request):
        qs = super().get_queryset(request).prefetch_related('options', 'survey_questions')
        if request.user.is_superuser:
            return qs
        return qs.filter(Q(is_public=True) | Q(created_by=request.user))
    
    # 自定义显示字段
    def text_preview(self, obj):
        """问题文本预览"""
        return obj.text[:80] + '...' if len(obj.text) > 80 else obj.text
    text_preview.short_description = '问题文本'
    
    def question_type_display(self, obj):
        """显示中文问题类型"""
        return obj.get_question_type_display()
    question_type_display.short_description = '问题类型'
    
    def option_count(self, obj):
        """选项数量"""
        return obj.options.count()
    option_count.short_description = '选项数'
    
    def survey_usage_count(self, obj):
        """统计问题在问卷中被使用的次数"""
        return obj.survey_questions.count()
    survey_usage_count.short_description = '使用次数'
    
    # 批量操作
    def make_public(self, request, queryset):
        """批量设为公开"""
        #  对数据库执行操作，update()方法会返回一个整数，表示受影响/被更新的记录行数
        updated = queryset.update(is_public=True)
        self.message_user(request, f'成功将 {updated} 个问题设为公开')
    make_public.short_description = '设为公开'
    
    def make_private(self, request, queryset):
        """批量设为私有"""
        updated = queryset.update(is_public=False)
        self.message_user(request, f'成功将 {updated} 个问题设为私有')
    make_private.short_description = '设为私有'
    
    def change_category(self, request, queryset):
        """批量更改分类"""
        # 1. 提取选中数据的主键（PK）列表
        selected_ids = list(queryset.values_list('pk', flat=True))   
        # 2. 只有选中了数据才执行跳转（避免空选时的无效操作）
        if selected_ids:
            # 3. 反向解析「批量改分类」的后台 URL
            url = reverse('admin:survey_question_change_category')            
            # 4. 拼接选中数据的 ID 到 URL 中（用逗号分隔多个 ID）
            url += f'?ids={",".join(map(str, selected_ids))}'            
            # 5. 重定向到批量改分类的页面
            return HttpResponseRedirect(url)

    # 6. 给 Admin 操作下拉框设置显示名称（中文友好）
    change_category.short_description = '更改分类'
    
    def export_questions(self, request, queryset):
        """导出选中的问题（CSV格式）"""
        return self._export_csv(queryset)
    export_questions.short_description = '导出问题为CSV'
    
    def export_questions_excel(self, request, queryset):
        """导出选中的问题（Excel格式）"""
        return self._export_excel(queryset)
    export_questions_excel.short_description = '导出问题为Excel'
    
    # 导出辅助方法
    def _export_csv(self, queryset):
        """导出为CSV格式"""
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = 'attachment; filename="questions_export.csv"'
        
        writer = csv.writer(response)
        # 写入表头
        writer.writerow(['ID', '问题文本', '问题类型', '分类', '创建者', '是否公开', '创建时间', '选项(格式: 标签1;标签2;标签3)'])
        
        # 写入数据
        for question in queryset.select_related('category', 'created_by').prefetch_related('options'):
            options_str = ''
            if question.question_type in ['single_choice', 'multiple_choice']:
                options = [option.label for option in question.options.order_by('order')]
                options_str = ';'.join(options)
            
            writer.writerow([
                question.id,
                question.text,
                question.get_question_type_display(),
                question.category.name if question.category else '未分类',
                question.created_by.username,
                '是' if question.is_public else '否',
                question.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                options_str
            ])
        
        return response
    
    def _export_excel(self, queryset):
        """导出为Excel格式"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
        except ImportError:
            return HttpResponse('请先安装 openpyxl 库：pip install openpyxl', status=500)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "问题导出"
        
        # 设置表头样式
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 写入表头
        headers = ['ID', '问题文本', '问题类型', '分类', '创建者', '是否公开', '创建时间', '选项(格式: 标签1;标签2;标签3)']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
        
        # 写入数据
        for row_idx, question in enumerate(queryset.select_related('category', 'created_by').prefetch_related('options'), 2):
            options_str = ''
            if question.question_type in ['single_choice', 'multiple_choice']:
                options = [option.label for option in question.options.order_by('order')]
                options_str = ';'.join(options)
            
            ws.cell(row=row_idx, column=1, value=question.id)
            ws.cell(row=row_idx, column=2, value=question.text)
            ws.cell(row=row_idx, column=3, value=question.get_question_type_display())
            ws.cell(row=row_idx, column=4, value=question.category.name if question.category else '未分类')
            ws.cell(row=row_idx, column=5, value=question.created_by.username)
            ws.cell(row=row_idx, column=6, value='是' if question.is_public else '否')
            ws.cell(row=row_idx, column=7, value=question.created_at.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row_idx, column=8, value=options_str)
        
        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="questions_export.xlsx"'
        wb.save(response)
        return response
    
    # 自定义URL和视图
    def get_urls(self):
        from django.urls import path
        urls = super().get_urls() # 获取父类 ModelAdmin 已有的所有 URL 模式
        custom_urls = [
            # 定义新的 URL 路径，并将它们映射到相应的视图函数
            path('change_category/', self.admin_site.admin_view(self.change_category_view), 
                name='survey_question_change_category'), # 批量更改分类页面
            path('import_questions/', self.admin_site.admin_view(self.import_questions_view), 
                name='survey_question_import_questions'), # 导入问题页面
            path('export_template/', self.admin_site.admin_view(self.export_template_view), 
                name='survey_question_export_template'), # 导出模板页面
        ]
        # 将自定义 URL 放在标准 URL 之前，这样可以覆盖掉可能存在的同名默认 URL
        # 这是一个常见的模式，确保自定义视图优先匹配
        return custom_urls + urls 
    
    def change_category_view(self, request):
        """批量更改分类的视图"""
        # 动态定义一个表单类，用于选择新分类和接收要修改的问题ID
        class CategoryChangeForm(forms.Form):
            category = forms.ModelChoiceField(
                queryset=Category.objects.all(), # 下拉框的数据源是所有 Category 对象
                label='新分类',
                empty_label="请选择分类",  # 允许用户不选择任何分类
                widget=forms.Select(attrs={
                    'style': 'height: 45px;',  # 增加高度
                    'class': 'custom-select',  # 添加自定义类，方便CSS控制
                })
            )
            ids = forms.CharField(widget=forms.HiddenInput()) # 存储被选中的问题 ID 字符串，用 HiddenInput 隐藏起来

        if request.method == 'POST':
            form = CategoryChangeForm(request.POST)
            if form.is_valid():
                category = form.cleaned_data['category'] # 获取用户选择的新分类对象
                ids = form.cleaned_data['ids'].split(',') # 获取隐藏字段中的 ID 字符串，并按逗号分割成列表
                # 执行批量更新操作，将指定 ID 的问题的 category 字段设置为新分类
                Question.objects.filter(id__in=ids).update(category=category) 
                # 在 Admin 界面显示一条成功消息
                self.message_user(request, f'成功更新 {len(ids)} 个问题的分类')
                # 重定向回问题列表页
                return redirect(reverse('admin:survey_question_changelist')) 
        else: # GET 请求，通常是用户点击按钮后第一次访问这个页面
            ids = request.GET.get('ids', '') # 从 URL 查询参数中获取被选中的问题 ID 字符串
            form = CategoryChangeForm(initial={'ids': ids}) # 初始化表单，将 ID 字符串填充到隐藏字段中
        
        # 渲染模板，传递表单和其他上下文数据
        return render(request, 'admin/survey/change_category.html', {
            'form': form,
            'title': '批量更改问题分类', # 页面标题
            'opts': self.model._meta, # 传递模型元数据，常用于模板中构建导航等
        })
    
    def import_questions_view(self, request):
        """导入问题的视图"""
        # 动态定义一个表单类，用于上传文件和设置公共属性
        class ImportQuestionsForm(forms.Form):
            file = forms.FileField(
                label='文件 (支持CSV和Excel格式)',
                help_text='请上传包含问题的CSV或Excel文件'
            )
            is_public = forms.BooleanField(
                label='设为公开', 
                initial=True, # 默认勾选
                required=False, # 可以不勾选
                help_text='导入的问题是否公开可见'
            )
            
            # 自定义文件验证逻辑
            def clean_file(self):
                file = self.cleaned_data['file']
                allowed_types = [
                    'text/csv', # CSV 文件 MIME 类型
                    'application/vnd.ms-excel', # .xls 文件 MIME 类型
                    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' # .xlsx 文件 MIME 类型
                ]
                allowed_extensions = ['.csv', '.xls', '.xlsx']
                
                # 检查 MIME 类型或文件扩展名
                if file.content_type not in allowed_types:
                    ext = os.path.splitext(file.name)[1].lower() # 获取文件扩展名
                    if ext not in allowed_extensions:
                        raise forms.ValidationError('只支持CSV和Excel文件格式')
                
                # 检查文件大小（限制为5MB）
                max_size = 5 * 1024 * 1024
                if file.size > max_size:
                    raise forms.ValidationError(f'文件大小不能超过{max_size//1024//1024}MB')
                
                return file # 返回清理后的文件对象

        if request.method == 'POST':
            form = ImportQuestionsForm(request.POST, request.FILES) # 处理包含文件上传的 POST 请求
            if form.is_valid():
                try:
                    file = request.FILES['file'] # 获取上传的文件
                    is_public = form.cleaned_data['is_public'] # 获取是否公开的设置
                    rows = self._read_import_file(file) # 调用辅助方法读取文件内容为字典列表
                    
                    created_count = 0 # 记录成功创建的数量
                    error_count = 0 # 记录失败的数量
                    error_messages = [] # 记录错误信息
                    
                    # 遍历每一行数据（跳过第一行表头）
                    for i, row in enumerate(rows, 2):  
                        try:
                            # --- 处理分类 ---
                            category_name = row.get('分类', '').strip() # 从行数据中获取分类名称
                            category = None
                            if category_name:
                                # 如果分类存在则获取，不存在则创建（根据名称和 slug）
                                category, _ = Category.objects.get_or_create(
                                    name=category_name, 
                                    defaults={'slug': category_name}
                                )
                            
                            # --- 处理问题类型 ---
                            question_type_input = row.get('问题类型', 'text').strip() # 获取原始输入
                            question_type = self._parse_question_type(question_type_input) # 转换为模型使用的枚举值
                            
                            # --- 创建问题 ---
                            question = Question.objects.create(
                                text=row.get('问题文本', '').strip(), # 获取问题文本
                                question_type=question_type, # 设置问题类型
                                category=category, # 设置分类（可能为 None）
                                created_by=request.user, # 设置创建者为当前登录的管理员
                                is_public=is_public # 设置是否公开（由表单决定）
                            )
                            created_count += 1
                            
                            # --- 处理选项 ---
                            options_str = self._get_options_string(row) # 获取选项字符串
                            # 如果问题是单选或多选，并且提供了选项，则创建选项
                            if options_str and question_type in ['single_choice', 'multiple_choice']:
                                self._create_options(question, options_str) # 调用辅助方法创建选项
                                    
                        except Exception as e:
                            error_count += 1
                            error_messages.append(f'第{i}行导入失败: {str(e)}') # 记录具体哪一行出错及原因
                    
                    # --- 显示最终结果 ---
                    if created_count > 0:
                        self.message_user(request, f'成功导入 {created_count} 个问题')
                    if error_count > 0:
                        # 只显示前5个错误，避免消息太长
                        self.message_user(request, f'有 {error_count} 个问题导入失败: {"; ".join(error_messages[:5])}', 'warning')
                    
                    return redirect(reverse('admin:survey_question_changelist')) # 成功后重定向到列表页
                        
                except Exception as e:
                    # 如果在文件读取或主处理过程中发生顶层异常，则将其作为表单非字段错误显示
                    form.add_error(None, f'文件处理失败: {str(e)}')
        else: # GET 请求，显示上传表单
            form = ImportQuestionsForm()
        
        # 渲染模板，传递表单
        return render(request, 'admin/survey/import_questions.html', {
            'form': form,
            'title': '导入问题',
            'opts': self.model._meta,
        })
    
    def _read_import_file(self, file):
        """读取导入文件"""
        # 文件路径加.的后缀名格式化成小写
        ext = os.path.splitext(file.name)[1].lower()
        
        if ext == '.csv':
            # 读取CSV文件
            # 读取文件内容，解码格式，并按行分割（处理包含换行符的字段）
            decoded_file = file.read().decode('utf-8').splitlines()
            # 创建CSV读取器，使用 DictReader 直接将每一行转换为字典（键为表头）
            reader = csv.DictReader(decoded_file)
            return list(reader)
        else:
            # 读取Excel文件
            try:
                from openpyxl import load_workbook
            except ImportError:
                raise Exception('请安装 openpyxl 库以支持Excel导入')
            
            wb = load_workbook(file)
            ws = wb.active
            
            # 获取表头
            header_row = []
            # 遍历第一行的单元格，将其值转换为字符串并去除首尾空格，添加到表头列表中
            for cell in ws[1]:
                header_row.append(str(cell.value).strip())
            
            # 读取数据行
            rows = []
            # 遍历从第二行开始的所有数据行（跳过表头）
            for row in ws.iter_rows(min_row=2, values_only=True):
                # 跳过空行（所有单元格都是 None）
                if all(cell is None for cell in row):
                    continue

                """
                遍历当前行的每个单元格值，将其转换为字符串并去除首尾空格，
                同时检查是否超出表头范围，避免索引错误。
                最后，将键值对添加到当前行的字典中。
                """
                row_dict = {}
                for i, cell_value in enumerate(row):
                    if i < len(header_row):
                        """
                        将当前单元格的值转换为字符串并去除首尾空格，
                        同时检查是否为 None（Excel 中的空单元格），
                        如果是则设为空字符串。
                        """
                        row_dict[header_row[i]] = str(cell_value).strip() if cell_value is not None else ''
                rows.append(row_dict)
            
            return rows
    
    def _parse_question_type(self, question_type):
        """解析问题类型"""
        question_type_mapping = {
            '文本题': 'text',
            '单选题': 'single_choice',
            '多选题': 'multiple_choice',
            '评分题': 'rating',
            '日期题': 'date',
            'text': 'text',
            'single_choice': 'single_choice',
            'multiple_choice': 'multiple_choice',
            'rating': 'rating',
            'date': 'date'
        }
        # 返回映射值，若未找到则默认返回 'text'
        return question_type_mapping.get(question_type, 'text')
    
    def _get_options_string(self, row):
        """
        从行数据中获取选项字符串

        这个函数的目的是在一个数据行（通常可能是一个字典，键是列名，值是单元格内容）
        中查找预定义的几个可能包含“选项”信息的字段名，并返回第一个找到的字段对应的值（去除首尾空白）。
        如果都找不到，则返回空字符串。
        """
        # 定义一个列表，包含了需要在 'row' 中查找的、可能代表“选项”的字段名称。
        # 这些名称看起来像是表格的表头。
        # 列表的顺序很重要，函数会按顺序查找，找到第一个就返回。
        possible_fields = ['选项', '选项(格式: 值|标签;值|标签)', 
                        '选项(格式: 值|标签)', '选项(格式: 标签1;标签2;标签3)']

        # 遍历上面定义的可能字段名列表
        for field in possible_fields:
            # 检查当前字段名 'field' 是否存在于输入的 'row' 数据中（例如，作为字典的键）
            if field in row:
                # 如果找到了，就获取该字段对应的值 row[field]，
                # 使用 .strip() 方法去除这个值开头和结尾可能存在的空格或换行符，
                # 然后立即返回这个处理后的字符串。
                return row[field].strip()
        
        # 如果遍历完整个 possible_fields 列表都没有在 'row' 中找到任何匹配的字段名，
        # 则执行到这里，返回一个空字符串 ''。
        return ''
    
    def _create_options(self, question, options_str):
        """创建问题选项"""
        options = options_str.split(';')
        for i, option in enumerate(options):
            if '|' in option:
                # 格式: 值|标签
                value, label = option.split('|', 1)
                value = value.strip()
                label = label.strip()
            else:
                # 格式: 标签
                label = option.strip()
                # 生成值：移除特殊字符，转为小写，空格替换为下划线
                value = re.sub(r'[^\w\s]', '', label)
                value = value.lower().replace(' ', '_')
                if not value:
                    value = f'option_{i+1}'
            
            if value and label:
                Option.objects.create(
                    question=question,
                    value=value,
                    label=label,
                    order=i
                )
    
    def export_template_view(self, request):
        """导出问题模板"""
        export_format = request.GET.get('format', 'csv')
        
        template_data = [
            ['问题文本', '问题类型', '分类', '是否必填', '选项(格式: 标签1;标签2;标签3)'],
            ['您对我们的产品整体满意度如何？', '评分题', '用户体验', '是', ''],
            ['您是通过什么渠道知道我们的？', '单选题', '用户信息', '是', '朋友推荐;广告;搜索引擎;社交媒体;其他'],
            ['您喜欢我们产品的哪些方面？', '多选题', '产品反馈', '否', '产品设计;产品质量;价格合理;客户服务;功能实用'],
            ['您有什么建议或意见？', '文本题', '产品反馈', '否', ''],
            ['您的生日是哪一天？', '日期题', '个人信息', '是', ''],
            ['您对我们的服务评价如何？', '评分题', '服务评价', '是', ''],
            ['您使用过我们的哪些产品？', '多选题', '产品使用', '是', '产品1;产品2;产品3;产品4'],
            ['您希望我们添加哪些功能？', '文本题', '产品建议', '否', ''],
            ['您是在哪里购买我们的产品的？', '单选题', '购买渠道', '否', '线上;线下;其他']
        ]
        
        if export_format == 'excel':
            return self._export_template_excel(template_data)
        else:
            return self._export_template_csv(template_data)
    
    def _export_template_csv(self, template_data):
        """导出CSV模板"""
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = 'attachment; filename="question_template.csv"'
        
        writer = csv.writer(response)
        for row in template_data:
            writer.writerow(row)
        
        return response
    
    def _export_template_excel(self, template_data):
        """导出Excel模板"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
        except ImportError:
            return HttpResponse('请先安装 openpyxl 库：pip install openpyxl', status=500)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "问题模板"
        
        # 设置表头样式
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 写入数据
        for row_idx, row_data in enumerate(template_data, 1):
            for col_idx, cell_value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                if row_idx == 1:
                    cell.font = header_font
                    cell.alignment = header_alignment
        
        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="question_template.xlsx"'
        wb.save(response)
        return response


@admin.register(Survey)
class SurveyAdmin(admin.ModelAdmin):
    """问卷管理"""
    list_display = ['title', 'created_by', 'created_at', 'is_active', 'response_count', 'view_statistics']
    list_filter = ['is_active', 'created_at']
    search_fields = ['title', 'description']
    readonly_fields = ['created_at', 'updated_at', 'statistics']
    inlines = [SurveyQuestionInline]
    list_select_related = ['created_by']
    
    # 优化查询集
    def get_queryset(self, request):
        return super().get_queryset(request).prefetch_related(
            'responses', 'survey_questions__question'
        ).annotate(response_count=Count('responses'))
    
    # 自定义字段
    def response_count(self, obj):
        return obj.response_count if hasattr(obj, 'response_count') else obj.responses.count()
    response_count.short_description = '回答数量'
    response_count.admin_order_field = 'response_count'
    
    def view_statistics(self, obj):
        """查看详细统计的链接"""
        url = reverse('admin:survey_statistics', args=[obj.pk])
        return format_html(
            '<a href="{}" class="button" target="_blank">查看统计</a>',
            url
        )
    view_statistics.short_description = '详细统计'
    
    def statistics(self, obj):
        """显示问卷统计信息"""
        total_responses = obj.responses.count()
        html = f"<h3>问卷统计</h3>"
        html += f"<p><strong>总回答数：</strong>{total_responses}</p>"
        
        survey_questions = obj.survey_questions.select_related('question').all()
        if survey_questions:
            html += "<h4>问题详情：</h4><ul>"
            for sq in survey_questions:
                answer_count = sq.question.answers.count()
                html += f"<li><strong>{sq.question.text}</strong> ({sq.question.get_question_type_display()})：{answer_count} 个回答</li>"
            html += "</ul>"
        
        return format_html(html)
    statistics.short_description = '统计信息'
    
    # 自定义URL和视图
    def get_urls(self):
        from django.urls import path
        urls = super().get_urls()
        custom_urls = [
            path('<uuid:pk>/statistics/', self.admin_site.admin_view(self.statistics_view), 
                 name='survey_statistics'),
        ]
        return custom_urls + urls
    
    def statistics_view(self, request, pk):
        """详细统计视图"""
        survey = self.get_object(request, pk)
        if not survey:
            self.message_user(request, '问卷不存在', 'error')
            return redirect(reverse('admin:survey_survey_changelist'))
        
        # 计算统计数据
        total_responses = survey.responses.count()
        survey_questions = survey.survey_questions.select_related('question').prefetch_related(
            'question__options', 'question__answers'
        ).all().order_by('order')
        
        questions_stats = []
        for sq in survey_questions:
            question = sq.question
            answers = question.answers.all()
            stats = self._calculate_question_stats(question, answers)
            questions_stats.append(stats)
        
        context = {
            **self.admin_site.each_context(request),
            'title': f'{survey.title} - 统计信息',
            'survey': survey,
            'total_responses': total_responses,
            'questions': questions_stats,
            'opts': self.model._meta,
        }
        
        return TemplateResponse(request, 'admin/survey/statistics.html', context)
    
    def _calculate_question_stats(self, question, answers):
        """计算问题统计数据"""
        stats = {
            'question': question,
            'answer_count': answers.count(),
            'type': question.question_type,
            'data': {},
            'options': []
        }
        
        if question.question_type in ['single_choice', 'multiple_choice']:
            # 选择题统计
            option_stats = {}
            for option in question.options.all():
                option_stats[option.value] = {
                    'label': option.label,
                    'count': 0,
                    'percentage': 0.0
                }
            
            for answer in answers:
                choices = answer.answer_choice
                if isinstance(choices, list):
                    for choice in choices:
                        if choice in option_stats:
                            option_stats[choice]['count'] += 1
                elif isinstance(choices, str) and choices in option_stats:
                    option_stats[choices]['count'] += 1
            
            total = answers.count() or 1
            for option_data in option_stats.values():
                option_data['percentage'] = (option_data['count'] / total) * 100
            
            stats['data'] = option_stats
            stats['options'] = list(question.options.all().values('value', 'label'))
            
        elif question.question_type == 'rating':
            # 评分题统计
            ratings = {}
            for i in range(1, 6):
                ratings[str(i)] = {'count': 0, 'percentage': 0.0}
            
            for answer in answers:
                rating = answer.answer_choice
                if isinstance(rating, list) and rating:
                    rating = rating[0]
                if isinstance(rating, str) and rating in ratings:
                    ratings[rating]['count'] += 1
            
            total = answers.count() or 1
            for rating_data in ratings.values():
                rating_data['percentage'] = (rating_data['count'] / total) * 100
            
            stats['data'] = ratings
            
        elif question.question_type == 'text':
            # 文本题统计
            text_answers = []
            for answer in answers[:10]:
                text = answer.answer_text[:100] + ('...' if len(answer.answer_text) > 100 else '')
                text_answers.append(text)
            stats['data'] = text_answers
        
        return stats


@admin.register(Response)
class ResponseAdmin(admin.ModelAdmin):
    """回答记录管理"""
    list_display = ['survey', 'submit_time', 'wechat_nickname', 'completion_time', 'answer_count']
    list_filter = ['submit_time', 'survey']
    search_fields = ['wechat_nickname', 'wechat_openid', 'survey__title']
    readonly_fields = ['submit_time']
    list_select_related = ['survey']
    
    def has_add_permission(self, request):
        return False
    
    def answer_count(self, obj):
        return obj.answers.count()
    answer_count.short_description = '答案数量'


@admin.register(Answer)
class AnswerAdmin(admin.ModelAdmin):
    """答案管理"""
    list_display = ['question_preview', 'survey_preview', 'response', 'answer_preview']
    list_filter = ['response__survey', 'question__question_type']
    search_fields = ['answer_text', 'question__text', 'response__wechat_nickname']
    list_select_related = ['question', 'response__survey']
    
    def question_preview(self, obj):
        return obj.question.text[:50] + '...' if len(obj.question.text) > 50 else obj.question.text
    question_preview.short_description = '问题'
    
    def survey_preview(self, obj):
        return obj.response.survey.title
    survey_preview.short_description = '问卷'
    
    def answer_preview(self, obj):
        """答案预览，处理各种类型的答案"""
        if obj.answer_text:
            return obj.answer_text[:50] + '...' if len(obj.answer_text) > 50 else obj.answer_text
        elif obj.answer_choice:
            if obj.question.question_type in ['single_choice', 'multiple_choice']:
                choices = obj.answer_choice if isinstance(obj.answer_choice, list) else [obj.answer_choice]
                
                # 创建选项映射
                option_map = {option.value: option.label for option in obj.question.options.all()}
                
                # 转换值为标签
                labels = [option_map.get(choice, choice) for choice in choices]
                return ', '.join(labels)
            return str(obj.answer_choice)
        return '-'
    answer_preview.short_description = '答案'


@admin.register(QRCode)
class QRCodeAdmin(admin.ModelAdmin):
    """二维码管理"""
    list_display = ['name', 'survey', 'short_code', 'scan_count', 'created_at', 'qr_code_preview']
    list_filter = ['survey', 'created_at']
    search_fields = ['name', 'short_code', 'survey__title']
    readonly_fields = ['scan_count', 'created_at', 'qr_code_preview', 'download_qrcode']
    list_select_related = ['survey']
    
    fieldsets = (
        (None, {
            'fields': ('survey', 'name', 'short_code')
        }),
        ('统计信息', {
            'fields': ('scan_count', 'created_at'),
            'classes': ('collapse',)
        }),
        ('二维码', {
            'fields': ('qr_code_preview', 'download_qrcode'),
        }),
    )
    
    def get_form(self, request, obj=None, **kwargs):
        """处理GET参数，自动填充survey字段"""
        form = super().get_form(request, obj, **kwargs)
        survey_id = request.GET.get('survey')
        if survey_id and not obj:
            try:
                from .models import Survey
                Survey.objects.get(pk=survey_id)  # 验证survey存在
                form.base_fields['survey'].initial = survey_id
                form.base_fields['survey'].widget.attrs['readonly'] = True
                form.base_fields['survey'].disabled = True
            except Exception:
                pass
        return form
    
    def save_model(self, request, obj, form, change):
        """保存模型时自动生成短代码"""
        if not change and request.GET.get('survey'):
            try:
                from .models import Survey
                survey = Survey.objects.get(pk=request.GET.get('survey'))
                obj.survey = survey
            except Survey.DoesNotExist:
                pass
        
        # 生成唯一的短代码
        if not obj.short_code:
            obj.short_code = self._generate_unique_short_code()
        
        super().save_model(request, obj, form, change)
    
    def _generate_unique_short_code(self, length=8):
        """生成唯一的短代码"""
        max_attempts = 10
        for _ in range(max_attempts):
            code = ''.join(random.choices(string.ascii_lowercase + string.digits, k=length))
            if not QRCode.objects.filter(short_code=code).exists():
                return code
        # 如果多次尝试失败，使用UUID
        return str(uuid.uuid4())[:8]
    
    def qr_code_preview(self, obj):
        """二维码预览"""
        if not obj.short_code:
            return '请先保存生成二维码'
        
        return format_html('''
            <div style="margin: 10px 0;">
                <strong>二维码预览：</strong><br>
                <img src="/qrcode/{}/image/" alt="二维码" style="width: 200px; height: 200px; margin: 10px 0; border: 1px solid #ddd; padding: 5px;"><br>
                <a href="/qrcode/{}/image/" target="_blank" style="margin-right: 10px;">查看大图</a>
            </div>
        ''', obj.short_code, obj.short_code)
    qr_code_preview.short_description = '二维码预览'
    
    def download_qrcode(self, obj):
        """下载二维码"""
        if not obj.short_code:
            return ''
        
        return format_html('''
            <div style="margin: 10px 0;">
                <a href="/qrcode/{}/image/" download="qrcode_{}_{}.png" class="button" style="background-color: #4CAF50; color: white; padding: 8px 16px; text-decoration: none; border-radius: 4px;">
                    📥 下载二维码
                </a>
            </div>
        ''', obj.short_code, obj.short_code, obj.name)
    download_qrcode.short_description = '下载'

