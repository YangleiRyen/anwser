# survey/admin/question_admin.py
from django.contrib import admin
from django.utils.html import format_html
from django.urls import reverse
from django.template.response import TemplateResponse
from django.http import HttpResponseRedirect, HttpResponse
from django.utils.translation import gettext_lazy as _
from django.db.models import Q, Count
from django import forms
from django.shortcuts import render, redirect
import csv
import os
import re
import random
import string
import uuid

from ..models import Survey, Question, Response, Answer, QRCode, Option, SurveyQuestion, Category


class OptionInline(admin.TabularInline):
    """问题选项内联表单"""
    model = Option
    extra = 1
    ordering = ['order']
    fields = ['value', 'label', 'order']
    verbose_name = '选项'
    verbose_name_plural = '选项'
    
    def formfield_for_dbfield(self, db_field, **kwargs):
        """优化表单字段显示"""
        field = super().formfield_for_dbfield(db_field, **kwargs)
        if db_field.name == 'label':
            field.widget.attrs['placeholder'] = '请输入选项显示文本'
        elif db_field.name == 'value':
            field.widget.attrs['placeholder'] = '自动生成，可自定义'
        return field


@admin.register(Question)
class QuestionAdmin(admin.ModelAdmin):
    """问题库管理"""
    list_display = ['text_preview', 'question_type_display', 'category', 'created_by', 
                    'is_public', 'created_at', 'option_count', 'survey_usage_count']
    list_filter = ['question_type', 'category', 'is_public', 'created_at']
    search_fields = ['text', 'category__name']
    list_select_related = ['category', 'created_by']
    ordering = ['category__name', '-created_at']
    inlines = [OptionInline]
    fields = ['text', 'question_type', 'category', 'created_by', 'is_public']
    change_list_template = 'admin/survey/question/change_list.html'
    actions = ['make_public', 'make_private', 'change_category', 'export_questions', 'export_questions_excel']
    list_editable = ['is_public', 'category']
    list_per_page = 20
    
    # 自定义表单字段
    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == 'created_by' and not request.user.is_superuser:
            kwargs['initial'] = request.user.id
            kwargs['disabled'] = True
        return super().formfield_for_foreignkey(db_field, request, **kwargs)
    
    # 自定义查询集
    def get_queryset(self, request):
        qs = super().get_queryset(request).prefetch_related('options', 'survey_questions')
        if request.user.is_superuser:
            return qs
        return qs.filter(Q(is_public=True) | Q(created_by=request.user))
    
    # 自定义显示字段
    def text_preview(self, obj):
        """问题文本预览"""
        return obj.text[:80] + '...' if len(obj.text) > 80 else obj.text
    text_preview.short_description = '问题文本'
    
    def question_type_display(self, obj):
        """显示中文问题类型"""
        return obj.get_question_type_display()
    question_type_display.short_description = '问题类型'
    
    def option_count(self, obj):
        """选项数量"""
        return obj.options.count()
    option_count.short_description = '选项数'
    
    def survey_usage_count(self, obj):
        """统计问题在问卷中被使用的次数"""
        return obj.survey_questions.count()
    survey_usage_count.short_description = '使用次数'
    
    # 批量操作
    def make_public(self, request, queryset):
        """批量设为公开"""
        updated = queryset.update(is_public=True)
        self.message_user(request, f'成功将 {updated} 个问题设为公开')
    make_public.short_description = '设为公开'
    
    def make_private(self, request, queryset):
        """批量设为私有"""
        updated = queryset.update(is_public=False)
        self.message_user(request, f'成功将 {updated} 个问题设为私有')
    make_private.short_description = '设为私有'
    
    def change_category(self, request, queryset):
        """批量更改分类"""
        selected_ids = list(queryset.values_list('pk', flat=True))
        if selected_ids:
            url = reverse('admin:survey_question_change_category')
            url += f'?ids={",".join(map(str, selected_ids))}'
            return HttpResponseRedirect(url)
    change_category.short_description = '更改分类'
    
    def export_questions(self, request, queryset):
        """导出选中的问题（CSV格式）"""
        return self._export_csv(queryset)
    export_questions.short_description = '导出问题为CSV'
    
    def export_questions_excel(self, request, queryset):
        """导出选中的问题（Excel格式）"""
        return self._export_excel(queryset)
    export_questions_excel.short_description = '导出问题为Excel'
    
    # 导出辅助方法
    def _export_csv(self, queryset):
        """导出为CSV格式"""
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = 'attachment; filename="questions_export.csv"'
        
        writer = csv.writer(response)
        # 写入表头
        writer.writerow(['ID', '问题文本', '问题类型', '分类', '创建者', '是否公开', '创建时间', '选项(格式: 标签1;标签2;标签3)'])
        
        # 写入数据
        for question in queryset.select_related('category', 'created_by').prefetch_related('options'):
            options_str = ''
            if question.question_type in ['single_choice', 'multiple_choice']:
                options = [option.label for option in question.options.order_by('order')]
                options_str = ';'.join(options)
            
            writer.writerow([
                question.id,
                question.text,
                question.get_question_type_display(),
                question.category.name if question.category else '未分类',
                question.created_by.username,
                '是' if question.is_public else '否',
                question.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                options_str
            ])
        
        return response
    
    def _export_excel(self, queryset):
        """导出为Excel格式"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
        except ImportError:
            return HttpResponse('请先安装 openpyxl 库：pip install openpyxl', status=500)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "问题导出"
        
        # 设置表头样式
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 写入表头
        headers = ['ID', '问题文本', '问题类型', '分类', '创建者', '是否公开', '创建时间', '选项(格式: 标签1;标签2;标签3)']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
        
        # 写入数据
        for row_idx, question in enumerate(queryset.select_related('category', 'created_by').prefetch_related('options'), 2):
            options_str = ''
            if question.question_type in ['single_choice', 'multiple_choice']:
                options = [option.label for option in question.options.order_by('order')]
                options_str = ';'.join(options)
            
            ws.cell(row=row_idx, column=1, value=question.id)
            ws.cell(row=row_idx, column=2, value=question.text)
            ws.cell(row=row_idx, column=3, value=question.get_question_type_display())
            ws.cell(row=row_idx, column=4, value=question.category.name if question.category else '未分类')
            ws.cell(row=row_idx, column=5, value=question.created_by.username)
            ws.cell(row=row_idx, column=6, value='是' if question.is_public else '否')
            ws.cell(row=row_idx, column=7, value=question.created_at.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row_idx, column=8, value=options_str)
        
        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="questions_export.xlsx"'
        wb.save(response)
        return response
    
    # 自定义URL和视图
    def get_urls(self):
        from django.urls import path
        urls = super().get_urls()
        custom_urls = [
            path('change_category/', self.admin_site.admin_view(self.change_category_view), 
                 name='survey_question_change_category'),
            path('import_questions/', self.admin_site.admin_view(self.import_questions_view), 
                 name='survey_question_import_questions'),
            path('export_template/', self.admin_site.admin_view(self.export_template_view), 
                 name='survey_question_export_template'),
        ]
        return custom_urls + urls
    
    def change_category_view(self, request):
        """批量更改分类的视图"""
        class CategoryChangeForm(forms.Form):
            category = forms.ModelChoiceField(
                queryset=Category.objects.all(), 
                label='新分类',
                empty_label="请选择分类"
            )
            ids = forms.CharField(widget=forms.HiddenInput())
        
        if request.method == 'POST':
            form = CategoryChangeForm(request.POST)
            if form.is_valid():
                category = form.cleaned_data['category']
                ids = form.cleaned_data['ids'].split(',')
                Question.objects.filter(id__in=ids).update(category=category)
                self.message_user(request, f'成功更新 {len(ids)} 个问题的分类')
                return redirect(reverse('admin:survey_question_changelist'))
        else:
            ids = request.GET.get('ids', '')
            form = CategoryChangeForm(initial={'ids': ids})
        
        return render(request, 'admin/survey/change_category.html', {
            'form': form,
            'title': '批量更改问题分类',
            'opts': self.model._meta,
        })
    
    def import_questions_view(self, request):
        """导入问题的视图"""
        class ImportQuestionsForm(forms.Form):
            file = forms.FileField(
                label='文件 (支持CSV和Excel格式)',
                help_text='请上传包含问题的CSV或Excel文件'
            )
            is_public = forms.BooleanField(
                label='设为公开', 
                initial=True, 
                required=False,
                help_text='导入的问题是否公开可见'
            )
            
            def clean_file(self):
                file = self.cleaned_data['file']
                allowed_types = [
                    'text/csv',
                    'application/vnd.ms-excel',
                    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                ]
                allowed_extensions = ['.csv', '.xls', '.xlsx']
                
                # 检查文件类型
                if file.content_type not in allowed_types:
                    ext = os.path.splitext(file.name)[1].lower()
                    if ext not in allowed_extensions:
                        raise forms.ValidationError('只支持CSV和Excel文件格式')
                
                # 检查文件大小（限制为5MB）
                max_size = 5 * 1024 * 1024
                if file.size > max_size:
                    raise forms.ValidationError(f'文件大小不能超过{max_size//1024//1024}MB')
                
                return file
        
        if request.method == 'POST':
            form = ImportQuestionsForm(request.POST, request.FILES)
            if form.is_valid():
                try:
                    file = request.FILES['file']
                    is_public = form.cleaned_data['is_public']
                    rows = self._read_import_file(file)
                    
                    created_count = 0
                    error_count = 0
                    error_messages = []
                    
                    for i, row in enumerate(rows, 2):  # 从第2行开始计数（表头是第1行）
                        try:
                            # 处理分类
                            category_name = row.get('分类', '').strip()
                            category = None
                            if category_name:
                                category, _ = Category.objects.get_or_create(
                                    name=category_name, 
                                    defaults={'slug': category_name}
                                )
                            
                            # 处理问题类型
                            question_type = self._parse_question_type(row.get('问题类型', 'text').strip())
                            
                            # 创建问题
                            question = Question.objects.create(
                                text=row.get('问题文本', '').strip(),
                                question_type=question_type,
                                category=category,
                                created_by=request.user,
                                is_public=is_public
                            )
                            created_count += 1
                            
                            # 处理选项
                            options_str = self._get_options_string(row)
                            if options_str and question_type in ['single_choice', 'multiple_choice']:
                                self._create_options(question, options_str)
                                
                        except Exception as e:
                            error_count += 1
                            error_messages.append(f'第{i}行导入失败: {str(e)}')
                    
                    # 显示结果消息
                    if created_count > 0:
                        self.message_user(request, f'成功导入 {created_count} 个问题')
                    if error_count > 0:
                        self.message_user(request, f'有 {error_count} 个问题导入失败: {"; ".join(error_messages[:5])}', 'warning')
                    
                    return redirect(reverse('admin:survey_question_changelist'))
                    
                except Exception as e:
                    form.add_error(None, f'文件处理失败: {str(e)}')
        else:
            form = ImportQuestionsForm()
        
        return render(request, 'admin/survey/import_questions.html', {
            'form': form,
            'title': '导入问题',
            'opts': self.model._meta,
        })
    
    def _read_import_file(self, file):
        """读取导入文件"""
        ext = os.path.splitext(file.name)[1].lower()
        
        if ext == '.csv':
            # 读取CSV文件
            decoded_file = file.read().decode('utf-8').splitlines()
            reader = csv.DictReader(decoded_file)
            return list(reader)
        else:
            # 读取Excel文件
            try:
                from openpyxl import load_workbook
            except ImportError:
                raise Exception('请安装 openpyxl 库以支持Excel导入')
            
            wb = load_workbook(file)
            ws = wb.active
            
            # 获取表头
            header_row = []
            for cell in ws[1]:
                header_row.append(str(cell.value).strip())
            
            # 读取数据行
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if all(cell is None for cell in row):
                    continue
                
                row_dict = {}
                for i, cell_value in enumerate(row):
                    if i < len(header_row):
                        row_dict[header_row[i]] = str(cell_value).strip() if cell_value is not None else ''
                rows.append(row_dict)
            
            return rows
    
    def _parse_question_type(self, question_type):
        """解析问题类型"""
        question_type_mapping = {
            '文本题': 'text',
            '单选题': 'single_choice',
            '多选题': 'multiple_choice',
            '评分题': 'rating',
            '日期题': 'date',
            'text': 'text',
            'single_choice': 'single_choice',
            'multiple_choice': 'multiple_choice',
            'rating': 'rating',
            'date': 'date'
        }
        
        return question_type_mapping.get(question_type, 'text')
    
    def _get_options_string(self, row):
        """从行数据中获取选项字符串"""
        possible_fields = ['选项', '选项(格式: 值|标签;值|标签)', 
                          '选项(格式: 值|标签)', '选项(格式: 标签1;标签2;标签3)']
        for field in possible_fields:
            if field in row:
                return row[field].strip()
        return ''
    
    def _create_options(self, question, options_str):
        """创建问题选项"""
        options = options_str.split(';')
        for i, option in enumerate(options):
            if '|' in option:
                # 格式: 值|标签
                value, label = option.split('|', 1)
                value = value.strip()
                label = label.strip()
            else:
                # 格式: 标签
                label = option.strip()
                # 生成值：移除特殊字符，转为小写，空格替换为下划线
                value = re.sub(r'[^\w\s]', '', label)
                value = value.lower().replace(' ', '_')
                if not value:
                    value = f'option_{i+1}'
            
            if value and label:
                Option.objects.create(
                    question=question,
                    value=value,
                    label=label,
                    order=i
                )
    
    def export_template_view(self, request):
        """导出问题模板"""
        export_format = request.GET.get('format', 'csv')
        
        template_data = [
            ['问题文本', '问题类型', '分类', '是否必填', '选项(格式: 标签1;标签2;标签3)'],
            ['您对我们的产品整体满意度如何？', '评分题', '用户体验', '是', ''],
            ['您是通过什么渠道知道我们的？', '单选题', '用户信息', '是', '朋友推荐;广告;搜索引擎;社交媒体;其他'],
            ['您喜欢我们产品的哪些方面？', '多选题', '产品反馈', '否', '产品设计;产品质量;价格合理;客户服务;功能实用'],
            ['您有什么建议或意见？', '文本题', '产品反馈', '否', ''],
            ['您的生日是哪一天？', '日期题', '个人信息', '是', ''],
            ['您对我们的服务评价如何？', '评分题', '服务评价', '是', ''],
            ['您使用过我们的哪些产品？', '多选题', '产品使用', '是', '产品1;产品2;产品3;产品4'],
            ['您希望我们添加哪些功能？', '文本题', '产品建议', '否', ''],
            ['您是在哪里购买我们的产品的？', '单选题', '购买渠道', '否', '线上;线下;其他']
        ]
        
        if export_format == 'excel':
            return self._export_template_excel(template_data)
        else:
            return self._export_template_csv(template_data)
    
    def _export_template_csv(self, template_data):
        """导出CSV模板"""
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = 'attachment; filename="question_template.csv"'
        
        writer = csv.writer(response)
        for row in template_data:
            writer.writerow(row)
        
        return response
    
    def _export_template_excel(self, template_data):
        """导出Excel模板"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
        except ImportError:
            return HttpResponse('请先安装 openpyxl 库：pip install openpyxl', status=500)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "问题模板"
        
        # 设置表头样式
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 写入数据
        for row_idx, row_data in enumerate(template_data, 1):
            for col_idx, cell_value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                if row_idx == 1:
                    cell.font = header_font
                    cell.alignment = header_alignment
        
        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="question_template.xlsx"'
        wb.save(response)
        return response
